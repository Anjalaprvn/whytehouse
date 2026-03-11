from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.text import slugify
from django.db import transaction
from django.urls import reverse
import random
from django.core.mail import send_mail
from django.contrib import messages
from datetime import datetime
from .models import BlogCategory
from django.http import JsonResponse


from .models import Account, Customer, Resort, Meal, Voucher, Invoice, Lead, Property, TravelPackage, Inquiry, Destination, Feedback
from .models import Employee,Blog,BlogImage

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# Create your views here.
def home(request):
    return redirect('login')



def login(request):
    if request.method == "POST":
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")

        user = User.objects.filter(email__iexact=email).first()

        if user and password and user.check_password(password):
            # Check if user has email configured
            if not user.email:
                return render(request, "admin/login.html", {
                    "error": "Email not configured for this account"
                })

            # Generate OTP
            otp = random.randint(100000, 999999)

            # Store in session
            request.session["admin_otp"] = str(otp)
            request.session["admin_user_id"] = user.id

            # Send OTP email
            try:
                send_mail(
                    subject="Admin Login OTP",
                    message=f"Your OTP for login is: {otp}",
                    from_email="whytehousee@gmail.com",
                    recipient_list=[user.email],
                    fail_silently=False,
                )
                messages.success(request, "OTP sent to your registered email.")
                return redirect("admin_panel:verify_otp")
            except Exception as e:
                return render(request, "admin/login.html", {
                    "error": f"Failed to send OTP: {str(e)}"
                })

        else:
            return render(request, "admin/login.html", {
                "error": "Invalid email or password"
            })

    return render(request, "admin/login.html")

def verify_otp(request):
    # Check if session has required data
    session_otp = request.session.get("admin_otp")
    user_id = request.session.get("admin_user_id")
    
    if not session_otp or not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect("admin_panel:login")
    
    if request.method == "POST":
        entered_otp = request.POST.get("otp", "").strip()

        if entered_otp == session_otp:
            try:
                user = User.objects.get(id=user_id)
                
                # Login user officially
                auth_login(request, user)
                
                # Clear OTP session but keep verification flag
                request.session.pop("admin_otp", None)
                request.session.pop("admin_user_id", None)
                # Set OTP verified flag
                request.session['admin_otp_verified'] = True
                
                messages.success(request, "Login successful!")
                return redirect("admin_panel:dashboard")
            except User.DoesNotExist:
                messages.error(request, "User not found. Please login again.")
                return redirect("admin_panel:login")
        else:
            return render(request, "admin/verify_otp.html", {
                "error": "Invalid OTP. Please try again."
            })

    return render(request, "admin/verify_otp.html")
def resend_otp(request):
    user_id = request.session.get('admin_user_id')
    
    if not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect('admin_panel:login')
    
    try:
        user = User.objects.get(id=user_id)
        if not user.email:
            messages.error(request, "Email not configured for this account.")
            return redirect('admin_panel:login')
        
        # Generate new OTP
        otp = random.randint(100000, 999999)
        request.session['admin_otp'] = str(otp)  # Store as string
        
        # Send OTP email
        try:
            send_mail(
                subject="Admin Login OTP - Resent",
                message=f"Your new OTP for login is: {otp}",
                from_email="whytehousee@gmail.com",
                recipient_list=[user.email],
                fail_silently=False,
            )
            messages.success(request, "OTP has been resent to your email.")
        except Exception as e:
            messages.error(request, f"Failed to resend OTP: {str(e)}")
            
    except User.DoesNotExist:
        messages.error(request, "User not found. Please login again.")
        return redirect('admin_panel:login')
    
    return redirect('admin_panel:verify_otp')

def forgot_password(request):
    return render(request, 'admin/forgotpassword.html')

def dashboard(request):
    # Check if user has completed OTP verification (custom session check)
    otp_verified = request.session.get('admin_otp_verified', False)
    
    if not otp_verified:
        messages.error(request, "Please login to access this page.")
        return redirect('admin_panel:login')
    
    from django.db.models import Sum, Count
    from datetime import datetime, timedelta
    
    # Stats
    total_vouchers = Voucher.objects.count()
    total_invoices = Invoice.objects.count()
    total_profit = Invoice.objects.aggregate(Sum('profit'))['profit__sum'] or 0
    new_leads = Lead.objects.filter(created_at__gte=datetime.now() - timedelta(days=30)).count()
    total_customers = Customer.objects.count()
    total_feedbacks = Feedback.objects.count()
    total_blogs = Blog.objects.count()
    international_packages = TravelPackage.objects.filter(category='International').count()
    domestic_packages = TravelPackage.objects.filter(category='Domestic').count()
    
    # Upcoming bookings (vouchers with future check-in dates)
    upcoming_bookings = Voucher.objects.filter(checkin_date__gte=datetime.now()).order_by('checkin_date')[:3]
    
    # Recent invoices
    recent_invoices = Invoice.objects.select_related('customer').order_by('-created_at')[:5]
    
    # Recent leads
    recent_leads = Lead.objects.order_by('-created_at')[:5]
    
    context = {
        'total_vouchers': total_vouchers,
        'total_invoices': total_invoices,
        'total_profit': total_profit,
        'new_leads': new_leads,
        'total_customers': total_customers,
        'total_feedbacks': total_feedbacks,
        'total_blogs': total_blogs,
        'international_packages': international_packages,
        'domestic_packages': domestic_packages,
        'upcoming_bookings': upcoming_bookings,
        'recent_invoices': recent_invoices,
        'recent_leads': recent_leads,
    }
    return render(request, 'admin/index.html', context)

def logout_view(request):
    from django.contrib.auth import logout
    request.session.pop('admin_otp_verified', None)
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('admin_panel:login')

# LEADS
def lead_management(request):
    enquiry_type = request.GET.get('type', '')
    source_filter = request.GET.get('source', '')
    new_leads = request.GET.get('new', '')
    search_query = request.GET.get('search', '').strip()
    
    # If General enquiry type is selected, redirect to customer inquiries page
    if enquiry_type == 'General':
        return redirect('admin_panel:customer_inquiries')
    
    leads = Lead.objects.all()
    
    if search_query:
        leads = leads.filter(
            Q(full_name__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(place__icontains=search_query)
        )
    
    if enquiry_type:
        leads = leads.filter(enquiry_type=enquiry_type)
    
    if source_filter:
        leads = leads.filter(source=source_filter)
    
    if new_leads == 'true':
        leads = leads.filter(is_viewed=False)
    
    leads = leads.order_by('-created_at')
    
    # Count Inquiries for general (not Leads)
    general_count = Inquiry.objects.count()
    international_count = Lead.objects.filter(enquiry_type='International').count()
    domestic_count = Lead.objects.filter(enquiry_type='Domestic').count()
    new_leads_count = Lead.objects.filter(is_viewed=False).count()
    
    context = {
        'leads': leads,
        'selected_type': enquiry_type,
        'selected_source': source_filter,
        'selected_new': new_leads,
        'search_query': search_query,
        'general_count': general_count,
        'international_count': international_count,
        'domestic_count': domestic_count,
        'new_leads_count': new_leads_count,
    }
    return render(request, 'admin/lead/lead.html', context)

def add_lead(request):
    if request.method == "POST":
        Lead.objects.create(
            full_name=request.POST.get('full_name'),
            mobile_number=request.POST.get('mobile_number'),
            place=request.POST.get('place'),
            source=request.POST.get('source'),
            enquiry_type=request.POST.get('enquiry_type', 'General'),
            remarks=request.POST.get('remarks')
        )
        messages.success(request, "Lead added successfully!")
        return redirect('admin_panel:leads')
    return render(request, 'admin/lead/lead_add.html')

def edit_lead(request, id):
    lead = get_object_or_404(Lead, id=id)
    if request.method == "POST":
        lead.full_name = request.POST.get('full_name')
        lead.mobile_number = request.POST.get('mobile_number')
        lead.place = request.POST.get('place')
        lead.source = request.POST.get('source')
        lead.enquiry_type = request.POST.get('enquiry_type', 'General')
        lead.status = request.POST.get('status', 'New')
        lead.remarks = request.POST.get('remarks')
        lead.save()
        messages.success(request, "Lead updated successfully!")
        return redirect('admin_panel:leads')
    return render(request, 'admin/lead/lead_edit.html', {'lead': lead})

def delete_lead(request, lead_id):
    if request.method != 'POST':
        return redirect('admin_panel:leads')
    
    lead = get_object_or_404(Lead, id=lead_id)
    lead.delete()
    messages.success(request, 'Lead deleted successfully!')
    return redirect('admin_panel:leads')

def view_lead(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    # Mark lead as viewed
    if not lead.is_viewed:
        lead.is_viewed = True
        lead.save()
    return render(request, 'admin/lead/lead_view.html', {'lead': lead})

# HOSPITALITY
def hospitality_management(request):
    search_query = request.GET.get('search', '').strip()
    properties = Property.objects.all()
    
    if search_query:
        properties = properties.filter(
            Q(name__icontains=search_query) |
            Q(property_type__icontains=search_query) |
            Q(location__icontains=search_query)
        )
    
    properties = properties.order_by("-created_at")
    return render(request, "admin/hospitality/hospitality_management.html", {"properties": properties, "search_query": search_query})

def add_property(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        prop_type = request.POST.get("property_type", "").strip()
        location = request.POST.get("location", "").strip()
        website = request.POST.get("website", "").strip()
        owner_name = request.POST.get("owner_name", "").strip()
        owner_contact = request.POST.get("owner_contact", "").strip()

        if not name or not prop_type or not location:
            messages.error(request, "Name, type and location are required.")
            return render(request, "admin/hospitality/hospitality_add.html")
        
        # Validate website URL
        if website:
            import re
            url_pattern = re.compile(r'^https?://[\w\-]+(\.[\w\-]+)+[/#?]?.*$')
            if not url_pattern.match(website):
                messages.error(request, "Invalid website URL format. Must start with http:// or https://")
                return render(request, "admin/hospitality/hospitality_add.html")
        
        # Validate owner contact number
        if owner_contact:
            if not owner_contact.isdigit() or len(owner_contact) < 10:
                messages.error(request, "Owner contact must be a valid phone number (minimum 10 digits).")
                return render(request, "admin/hospitality/hospitality_add.html")
        
        new_amenities = request.POST.getlist("new_amenities[]")

        
        amenities_text = ", ".join([a.strip() for a in new_amenities if a.strip()])

        Property.objects.create(
            name=name,
            property_type=prop_type,
            location=location,
            website=website or None,
            address=request.POST.get("address"),
            summary=request.POST.get("summary"),
            owner_name=owner_name or None,
            owner_contact=owner_contact or None,
            amenities=amenities_text,
            image=request.FILES.get("image"),
        )

        messages.success(request, "Property added successfully!")
        return redirect("admin_panel:admin_hospitality")

    return render(request, "admin/hospitality/hospitality_add.html")

def edit_property(request, property_id):
    prop = get_object_or_404(Property, id=property_id)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        prop_type = request.POST.get("property_type", "").strip()
        location = request.POST.get("location", "").strip()
        if not name or not prop_type or not location:
            messages.error(request, "Name, type and location are required.")
            return render(
                request,
                "admin/hospitality/hospitality_edit.html",
                {"property": prop}
            )

        prop.name = name
        prop.property_type = prop_type
        prop.location = location
        prop.website = request.POST.get("website") or None
        prop.address = request.POST.get("address")
        prop.summary = request.POST.get("summary")
        prop.owner_name = request.POST.get("owner_name") or None
        prop.owner_contact = request.POST.get("owner_contact") or None

        # Get existing amenities from checkboxes
        selected_amenities = request.POST.getlist("amenities")
        # Get new amenity from text input
        new_amenity = request.POST.get("new_amenity", "").strip()
        
        # Combine all amenities
        all_amenities = list(selected_amenities)
        if new_amenity:
            all_amenities.append(new_amenity)
        
        prop.amenities = ", ".join([a.strip() for a in all_amenities if a.strip()])

        if request.FILES.get("image"):
            prop.image = request.FILES.get("image")

        prop.save()

        messages.success(request, "Property updated successfully!")
        return redirect("admin_panel:admin_hospitality")

    return render(
        request,
        "admin/hospitality/hospitality_edit.html",
        {"property": prop}
    )

def delete_property(request, property_id):
    prop = get_object_or_404(Property, id=property_id)
    prop.delete()
    messages.success(request, "Property deleted successfully!")
    return redirect("admin_panel:admin_hospitality")

def toggle_property_status(request, property_id):
    if request.method != 'POST':
        return redirect('admin_panel:admin_hospitality')
    
    prop = get_object_or_404(Property, id=property_id)
    prop.is_active = not prop.is_active
    prop.save()
    
    status_text = "enabled" if prop.is_active else "disabled"
    messages.success(request, f"Property '{prop.name}' has been {status_text} successfully!")
    return redirect('admin_panel:admin_hospitality')

def view_property(request, property_id):
    prop = get_object_or_404(Property, id=property_id)
    return render(request, "admin/hospitality/hospitality_view.html", {"property": prop})

# TRAVEL PACKAGES
def travel_packages(request):
    category = request.GET.get('cat', 'Domestic')
    destination_id = request.GET.get('dest')
    search_query = request.GET.get('search', '').strip()
    
    # Get destinations for the selected category
    destinations = Destination.objects.filter(category=category).order_by('name')
    
    # Get packages based on category and destination
    packages = TravelPackage.objects.filter(category=category)
    
    if search_query:
        packages = packages.filter(
            Q(name__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(country__icontains=search_query)
        )
    
    if destination_id:
        packages = packages.filter(destination_id=destination_id)
    packages = packages.order_by('-created_at')
    
    # Count packages by destination
    destination_counts = {}
    for dest in destinations:
        destination_counts[dest.id] = TravelPackage.objects.filter(category=category, destination=dest).count()
    
    domestic_count = TravelPackage.objects.filter(category='Domestic').count()
    international_count = TravelPackage.objects.filter(category='International').count()
    
    # Get selected destination object
    selected_destination_obj = None
    if destination_id:
        try:
            selected_destination_obj = Destination.objects.get(id=destination_id)
        except Destination.DoesNotExist:
            pass
    
    context = {
        'packages': packages,
        'destinations': destinations,
        'destination_counts': destination_counts,
        'selected_category': category,
        'selected_destination': int(destination_id) if destination_id else None,
        'selected_destination_obj': selected_destination_obj,
        'domestic_count': domestic_count,
        'international_count': international_count,
        'search_query': search_query,
    }
    return render(request, 'admin/packages/travel_packages.html', context)

def travel_package_add(request):
    # Get category and destination from URL parameters
    default_category = request.GET.get('category', 'Domestic')
    destination_id = request.GET.get('destination')
    
    # Get the selected destination object
    selected_destination_obj = None
    if destination_id:
        try:
            selected_destination_obj = Destination.objects.get(id=destination_id)
        except Destination.DoesNotExist:
            pass
    
    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        price = request.POST.get('price', '').strip()
        duration = request.POST.get('duration', '').strip()
        location = request.POST.get('location', '').strip()
        description = request.POST.get('description', '').strip()
        
        # Validation
        if not name or not price or not duration or not location:
            messages.error(request, "Name, price, duration, and location are required.")
            return render(request, 'admin/packages/travel_package_add.html', {
                'default_category': default_category,
                'selected_destination_id': int(destination_id) if destination_id else None,
                'selected_destination_obj': selected_destination_obj,
            })
        
        # Validate price
        try:
            price_val = float(price)
            if price_val < 0:
                messages.error(request, "Price must be a positive number.")
                return render(request, 'admin/packages/travel_package_add.html', {
                    'default_category': default_category,
                    'selected_destination_id': int(destination_id) if destination_id else None,
                    'selected_destination_obj': selected_destination_obj,
                })
        except ValueError:
            messages.error(request, "Invalid price format.")
            return render(request, 'admin/packages/travel_package_add.html', {
                'default_category': default_category,
                'selected_destination_id': int(destination_id) if destination_id else None,
                'selected_destination_obj': selected_destination_obj,
            })
        
        # Validate duration format (e.g., "3 Days 2 Nights")
        if not any(word in duration.lower() for word in ['day', 'night', 'hour']):
            messages.error(request, "Duration must include 'day', 'night', or 'hour'.")
            return render(request, 'admin/packages/travel_package_add.html', {
                'default_category': default_category,
                'selected_destination_id': int(destination_id) if destination_id else None,
                'selected_destination_obj': selected_destination_obj,
            })
        
        destination_id = request.POST.get('destination')
        destination = None
        if destination_id:
            try:
                destination = Destination.objects.get(id=destination_id)
            except Destination.DoesNotExist:
                pass
        
        category = request.POST.get('category')
        TravelPackage.objects.create(
            name=name,
            category=category,
            destination=destination,
            price=price,
            duration=duration,
            location=location,
            country=request.POST.get('country'),
            description=description,
            itinerary='\n'.join(request.POST.getlist('itinerary[]')),
            inclusions='\n'.join(request.POST.getlist('inclusions[]')),
            exclusions='\n'.join(request.POST.getlist('exclusions[]')),
            meta_title=request.POST.get('meta_title'),
            meta_description=request.POST.get('meta_description'),
            active=request.POST.get('active') == 'on',
            image=request.FILES.get('image')
        )
        messages.success(request, "Package added successfully!")
        
        # Redirect back to the same category and destination
        url = reverse('admin_panel:travel_packages')
        if destination:
            return redirect(f'{url}?cat={category}&dest={destination.id}')
        else:
            return redirect(f'{url}?cat={category}')
    
    context = {
        'default_category': default_category,
        'selected_destination_id': int(destination_id) if destination_id else None,
        'selected_destination_obj': selected_destination_obj,
    }
    return render(request, 'admin/packages/travel_package_add.html', context)

def travel_package_edit(request, package_id):
    package = get_object_or_404(TravelPackage, id=package_id)
    
    # Get all destinations
    destinations = Destination.objects.all().order_by('category', 'name')
    
    if request.method == "POST":
        destination_id = request.POST.get('destination')
        destination = None
        if destination_id:
            try:
                destination = Destination.objects.get(id=destination_id)
            except Destination.DoesNotExist:
                pass
        
        # Get the category from POST
        new_category = request.POST.get('category')
        
        package.name = request.POST.get('name')
        package.category = new_category
        package.destination = destination
        package.price = request.POST.get('price')
        package.duration = request.POST.get('duration')
        package.location = request.POST.get('location')
        package.country = request.POST.get('country')
        package.description = request.POST.get('description')
        package.itinerary = request.POST.get('itinerary')
        package.inclusions = request.POST.get('inclusions')
        package.exclusions = request.POST.get('exclusions')
        package.meta_title = request.POST.get('meta_title')
        package.meta_description = request.POST.get('meta_description')
        package.active = request.POST.get('active') == 'on'
        if request.FILES.get('image'):
            package.image = request.FILES.get('image')
        if request.FILES.get('story_main_image'):
            package.story_main_image = request.FILES.get('story_main_image')
        if request.FILES.get('story_side_image1'):
            package.story_side_image1 = request.FILES.get('story_side_image1')
        if request.FILES.get('story_side_image2'):
            package.story_side_image2 = request.FILES.get('story_side_image2')
        package.save()
        messages.success(request, "Package updated successfully!")
        
        # Redirect back to the same category and destination using the NEW category
        url = reverse('admin_panel:travel_packages')
        if destination:
            return redirect(f'{url}?cat={new_category}&dest={destination.id}')
        else:
            return redirect(f'{url}?cat={new_category}')
    
    context = {
        'package': package,
        'destinations': destinations,
    }
    return render(request, 'admin/packages/travel_package_edit.html', context)

def travel_package_delete(request, package_id):
    package = get_object_or_404(TravelPackage, id=package_id)
    category = package.category
    destination_id = package.destination_id if package.destination else None
    package.delete()
    messages.success(request, "Package deleted successfully!")
    
    # Redirect back to the same category and destination
    url = reverse('admin_panel:travel_packages')
    if destination_id:
        return redirect(f'{url}?cat={category}&dest={destination_id}')
    else:
        return redirect(f'{url}?cat={category}')

def travel_package_view(request, package_id):
    package = get_object_or_404(TravelPackage, id=package_id)
    return render(request, 'admin/packages/travel_package_view.html', {'package': package})

def toggle_package_status(request, package_id):
    """Toggle active/inactive status of a package"""
    if request.method == 'POST':
        try:
            package = get_object_or_404(TravelPackage, id=package_id)
            package.active = not package.active
            package.save()
            
            # Check if it's an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'active': package.active})
            else:
                # Regular form submission - redirect back
                messages.success(request, f'Package status updated to {"Active" if package.active else "Inactive"}')
                return redirect('admin_panel:travel_packages')
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            else:
                messages.error(request, f'Error updating status: {str(e)}')
                return redirect('admin_panel:travel_packages')
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# CUSTOMER INQUIRIES
def customer_inquiries(request):
    status_filter = request.GET.get('status')
    
    # Get all inquiries
    inquiries = Inquiry.objects.all()
    
    # Apply status filter if provided
    if status_filter:
        inquiries = inquiries.filter(status=status_filter)
    
    inquiries = inquiries.order_by('-created_at')
    
    # Count by status (from all inquiries, not just filtered)
    all_inquiries = Inquiry.objects.all()
    new_count = all_inquiries.filter(status='New').count()
    contacted_count = all_inquiries.filter(status='Contacted').count()
    converted_count = all_inquiries.filter(status='Converted').count()
    junk_count = all_inquiries.filter(status='Junk').count()
    
    print(f"DEBUG customer_inquiries: Total={all_inquiries.count()}, New={new_count}, Contacted={contacted_count}, Converted={converted_count}, Junk={junk_count}")
    
    context = {
        'inquiries': inquiries,
        'total_count': all_inquiries.count(),
        'new_count': new_count,
        'contacted_count': contacted_count,
        'converted_count': converted_count,
        'junk_count': junk_count,
    }
    return render(request, 'admin/enquiry/customer_inquiries.html', context)

def view_inquiry(request, inquiry_id):
    inquiry = get_object_or_404(Inquiry, id=inquiry_id)
    return render(request, 'admin/enquiry/customer_inquiry_view.html', {'inquiry': inquiry})

def update_inquiry_status(request, inquiry_id):
    if request.method == 'POST':
        new_status = request.POST.get('status')
        print(f"DEBUG: Updating inquiry_id={inquiry_id} to status={new_status}")
        
        if new_status in ['New', 'Contacted', 'Converted', 'Junk']:
            inquiry = get_object_or_404(Inquiry, id=inquiry_id)
            print(f"DEBUG: Found Inquiry (ID={inquiry.id}, Name={inquiry.name}), updating status from {inquiry.status} to {new_status}")
            inquiry.status = new_status
            inquiry.save()
            print(f"DEBUG: Inquiry saved, new status is {inquiry.status}")
            messages.success(request, f'Inquiry status updated to {new_status}')
            return redirect('admin_panel:view_inquiry', inquiry_id=inquiry_id)
    
    return redirect('admin_panel:customer_inquiries')

def blog_list(request):
    return render(request, 'admin/blog/blog_list.html')

#EMPLOYEE-----------------------------------------------------------------

def employee_list(request):
    employees = Employee.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(position__icontains=search_query) |
            Q(department__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        employees = employees.filter(status=status_filter)
    
    department_filter = request.GET.get('department', '')
    if department_filter:
        employees = employees.filter(department=department_filter)
    
    departments = Employee.objects.values_list('department', flat=True).distinct().order_by('department')
    department_count = departments.count()
    
    active_count = employees.filter(status='active').count()
    inactive_count = employees.filter(status='inactive').count()
    
    context = {
        'employees': employees,
        'search_query': search_query,
        'status_filter': status_filter,
        'department_filter': department_filter,
        'departments': departments,
        'active_count': active_count,
        'department_count': department_count,
        'now': datetime.now().strftime('%B %d, %Y')
    }
    return render(request, "admin/employee/employee.html", context)

def add_employee(request):
    if request.method == 'POST':
        try:
            # basic required validation
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            phone = request.POST.get('phone', '').strip()
            if not name or not email or not phone:
                messages.error(request, "Name, email and phone are required.")
                return render(request, "admin/employee/add_employee.html")

            employee = Employee.objects.create(
                name=name,
                email=email,
                phone=phone,
                role=request.POST.get('role', ''),
                department=request.POST.get('department', ''),
                join_date=request.POST.get('join_date') or None,
                salary=request.POST.get('salary') or None,
                status=request.POST.get('status', 'Active'),
            )
            
            if request.FILES.get('profile_picture'):
                employee.profile_picture = request.FILES['profile_picture']
                employee.save()
            
            messages.success(request, f'Employee {employee.name} added successfully!')
            return redirect('employee:employee_list')
        except Exception as e:
            messages.error(request, f'Error adding employee: {str(e)}')
    return render(request, "admin/employee/add_employee.html")
           
    

def edit_employee(request, pk):
    try:
        employee = Employee.objects.get(pk=pk)
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found.')
        return redirect('employee:employee_list')
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            phone = request.POST.get('phone', '').strip()
            if not name or not email or not phone:
                messages.error(request, "Name, email and phone are required.")
                return render(request, "admin/employee/edit_employee.html", {'employee': employee})

            employee.name = name
            employee.email = email
            employee.phone = phone
            employee.role = request.POST.get('role', '')
            employee.department = request.POST.get('department', '')
            employee.join_date = request.POST.get('join_date') or None
            employee.salary = request.POST.get('salary') or None
            employee.status = request.POST.get('status', 'Active')
            
            if request.FILES.get('profile_picture'):
                employee.profile_picture = request.FILES['profile_picture']
            
            employee.save()
            
            messages.success(request, f'Employee {employee.name} updated successfully!')
            return redirect('employee:employee_list')
        except Exception as e:
            messages.error(request, f'Error updating employee: {str(e)}')
    
    return render(request, "admin/employee/edit_employee.html", {'employee': employee})

def delete_employee(request, pk):
    if request.method != 'POST':
        return redirect('employee:employee_list')
    
    try:
        employee = Employee.objects.get(pk=pk)
        employee_name = employee.name
        employee.delete()
        messages.success(request, f'Employee {employee_name} deleted successfully.')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found.')
    except Exception as e:
        messages.error(request, f'Error deleting employee: {str(e)}')
    
    return redirect('employee:employee_list')

def view_employee(request, pk):
    try:
        employee = Employee.objects.get(pk=pk)
        return render(request, "admin/employee/view_employee.html", {'employee': employee})
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found.')
        return redirect('employee:employee_list')
#SALES--------------------------------------------------------------------
            #ACCOUNT

def account_list(request):
    search_query = request.GET.get('search', '').strip()
    accounts = Account.objects.all()
    
    if search_query:
        accounts = accounts.filter(
            Q(account_name__icontains=search_query) |
            Q(account_number__icontains=search_query) |
            Q(bank_name__icontains=search_query) |
            Q(ifsc_code__icontains=search_query)
        )
    
    return render(request, 'admin/sales/account/account.html', {
        'accounts': accounts,
        'search_query': search_query
    })

def add_account(request):
    if request.method == 'POST':
        try:
            # Get form data
            account_name = request.POST.get('account_name', '').strip()
            account_number = request.POST.get('account_number', '').strip()
            bank_name = request.POST.get('bank_name', '').strip()
            ifsc_code = request.POST.get('ifsc_code', '').strip()
            account_type = request.POST.get('account_type', 'current')
            
            # Validation
            if not account_name or not account_number or not bank_name or not ifsc_code:
                messages.error(request, 'All fields are required.')
                return render(request, 'admin/sales/account/add_account.html')
            
            # Validate account number (digits only, 9-18 characters)
            if not account_number.isdigit() or len(account_number) < 9 or len(account_number) > 18:
                messages.error(request, 'Account number must be 9-18 digits.')
                return render(request, 'admin/sales/account/add_account.html')
            
            # Validate IFSC code format (11 characters, alphanumeric)
            import re
            if not re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', ifsc_code.upper()):
                messages.error(request, 'Invalid IFSC code format. Must be 11 characters (e.g., SBIN0001234).')
                return render(request, 'admin/sales/account/add_account.html')
            
            # Check for duplicate account number
            if Account.objects.filter(account_number=account_number).exists():
                messages.error(request, 'Account number already exists.')
                return render(request, 'admin/sales/account/add_account.html')
            
            # Create account
            account = Account.objects.create(
                account_name=account_name,
                account_number=account_number,
                bank_name=bank_name,
                ifsc_code=ifsc_code.upper(),
                account_type=account_type
            )
            
            messages.success(request, f"Account '{account_name}' added successfully!")
            return redirect('sales:account_list')
            
        except Exception as e:
            messages.error(request, f'Error adding account: {str(e)}')
            return render(request, 'admin/sales/account/add_account.html')
    
    return render(request, 'admin/sales/account/add_account.html')

def view_account(request, account_id):
    try:
        account = Account.objects.get(id=account_id)
        return render(request, 'admin/sales/account/view_account.html', {
            'account': account
        })
    except Account.DoesNotExist:
        messages.error(request, "Account not found.")
        return redirect("sales:account_list")


def edit_account(request, account_id):
    try:
        account = Account.objects.get(id=account_id)
    except Account.DoesNotExist:
        messages.error(request, "Account not found.")
        return redirect("sales:account_list")
    
    if request.method == "POST":
        try:
            # Get form data
            account_name = request.POST.get('account_name', '').strip()
            account_number = request.POST.get('account_number', '').strip()
            bank_name = request.POST.get('bank_name', '').strip()
            ifsc_code = request.POST.get('ifsc_code', '').strip()
            account_type = request.POST.get('account_type', '').strip()
            
            # Validation
            if not account_name or not account_number or not bank_name:
                messages.error(request, 'Account name, number, and bank name are required.')
                return render(request, 'admin/sales/account/edit_account.html', {
                    'account': account
                })
            
            # Update account
            account.account_name = account_name
            account.account_number = account_number
            account.bank_name = bank_name
            account.ifsc_code = ifsc_code
            account.account_type = account_type
            
            account.save()
            
            messages.success(request, f"Account '{account_name}' has been updated successfully!")
            return redirect("sales:account_list")
            
        except Exception as e:
            return render(request, 'admin/sales/account/edit_account.html', {
                'account': account,
                'error': f'Error updating account: {str(e)}'
            })
    
    return render(request, 'admin/sales/account/edit_account.html', {
        'account': account
    })


def delete_account(request, account_id):
    if request.method != 'POST':
        return redirect('sales:account_list')
    
    try:
        account = Account.objects.get(id=account_id)
        account_name = account.account_name
        account.delete()
        messages.success(request, f"Account '{account_name}' has been deleted successfully!")
    except Account.DoesNotExist:
        messages.error(request, "Account not found.")
    except Exception as e:
        messages.error(request, f"Error deleting account: {str(e)}")
    
    return redirect("sales:account_list")

        #CUSTOMER

def customer_list(request):
    search_query = request.GET.get('search', '').strip()
    customers = Customer.objects.all()
    
    if search_query:
        customers = customers.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(display_name__icontains=search_query) |
            Q(contact_number__icontains=search_query) |
            Q(place__icontains=search_query)
        )
    
    print(f"=== CUSTOMER LIST DEBUG ===")
    print(f"Total customers in database: {customers.count()}")
    for customer in customers:
        print(f"  Customer {customer.id}: {customer.display_name} - {customer.contact_number}")
    return render(request, "admin/sales/customer/customer.html", {
        "customers": customers,
        "search_query": search_query
    })

def add_customer(request):
    context = {"form": {}}

    if request.method == "POST":
        try:
            same_as_whatsapp = request.POST.get("same_as_whatsapp") == "on"

            customer_type = (request.POST.get("customer_type") or "Individual").strip()
            salutation = (request.POST.get("salutation") or "").strip()
            first_name = (request.POST.get("first_name") or "").strip()
            last_name = (request.POST.get("last_name") or "").strip()
            display_name = (request.POST.get("display_name") or "").strip()
            place = (request.POST.get("place") or "").strip()
            contact_number = (request.POST.get("contact_number") or "").strip()
            whatsapp_number = (request.POST.get("whatsapp_number") or "").strip()
            work_number = (request.POST.get("work_number") or "").strip()
            gst_number = (request.POST.get("gst_number") or "").strip()

            # Keep values in template if error
            context["form"] = {
                "customer_type": customer_type,
                "salutation": salutation,
                "first_name": first_name,
                "last_name": last_name,
                "display_name": display_name,
                "place": place,
                "contact_number": contact_number,
                "same_as_whatsapp": same_as_whatsapp,
                "whatsapp_number": whatsapp_number,
                "work_number": work_number,
                "gst_number": gst_number,
            }

            # Required validation
            if not first_name or not display_name or not contact_number:
                context["error"] = "First Name, Display Name, and Contact Number are required."
                return render(request, "admin/sales/customer/add_customer.html", context)
            
            # First name validation
            if len(first_name) < 2 or len(first_name) > 50:
                context["error"] = "First Name must be between 2 and 50 characters."
                return render(request, "admin/sales/customer/add_customer.html", context)
            
            # Display name validation
            if len(display_name) < 2 or len(display_name) > 100:
                context["error"] = "Display Name must be between 2 and 100 characters."
                return render(request, "admin/sales/customer/add_customer.html", context)
            
            # Contact number validation
            import re
            if not contact_number.isdigit():
                context["error"] = "Contact Number must contain only digits."
                return render(request, "admin/sales/customer/add_customer.html", context)
            
            if len(contact_number) < 10 or len(contact_number) > 15:
                context["error"] = "Contact Number must be between 10 and 15 digits."
                return render(request, "admin/sales/customer/add_customer.html", context)

            # NO-JS WhatsApp logic
            if same_as_whatsapp:
                whatsapp_number = contact_number

            # IMPORTANT fallback:
            # If whatsapp_number is empty, set it to contact_number (prevents NOT NULL issues)
            if not whatsapp_number:
                whatsapp_number = contact_number
            
            # WhatsApp number validation if different from contact
            if whatsapp_number != contact_number:
                if not whatsapp_number.isdigit():
                    context["error"] = "WhatsApp Number must contain only digits."
                    return render(request, "admin/sales/customer/add_customer.html", context)
                
                if len(whatsapp_number) < 10 or len(whatsapp_number) > 15:
                    context["error"] = "WhatsApp Number must be between 10 and 15 digits."
                    return render(request, "admin/sales/customer/add_customer.html", context)
            
            # Work number validation (optional)
            if work_number:
                if not work_number.isdigit():
                    context["error"] = "Work Number must contain only digits."
                    return render(request, "admin/sales/customer/add_customer.html", context)
                
                if len(work_number) < 10 or len(work_number) > 15:
                    context["error"] = "Work Number must be between 10 and 15 digits."
                    return render(request, "admin/sales/customer/add_customer.html", context)
            
            # GST number validation (optional)
            if gst_number:
                gst_pattern = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'
                if not re.match(gst_pattern, gst_number.upper()):
                    context["error"] = "Invalid GST Number format. Must be 15 characters (e.g., 22AAAAA0000A1Z5)."
                    return render(request, "admin/sales/customer/add_customer.html", context)
                gst_number = gst_number.upper()
            
            # Place validation (optional)
            if place and len(place) > 100:
                context["error"] = "Place must not exceed 100 characters."
                return render(request, "admin/sales/customer/add_customer.html", context)

            # Optional: prevent duplicates if contact_number is unique
            if Customer.objects.filter(contact_number=contact_number).exists():
                context["error"] = "This Contact Number already exists."
                return render(request, "admin/sales/customer/add_customer.html", context)

            Customer.objects.create(
                customer_type=customer_type,
                salutation=salutation,
                first_name=first_name,
                last_name=last_name,
                display_name=display_name,
                place=place,
                contact_number=contact_number,
                same_as_whatsapp=same_as_whatsapp,
                whatsapp_number=whatsapp_number,
                work_number=work_number,
                gst_number=gst_number,
            )

            messages.success(request, f"Customer '{display_name}' added successfully!")
            return redirect("sales:customer_list")

        except IntegrityError as e:
            print("INTEGRITY ERROR:", str(e))
            context["error"] = f"Database error: {str(e)}"
            return render(request, "admin/sales/customer/add_customer.html", context)

        except Exception as e:
            import traceback
            traceback.print_exc()
            context["error"] = f"Error saving customer: {str(e)}"
            return render(request, "admin/sales/customer/add_customer.html", context)

    return render(request, "admin/sales/customer/add_customer.html", context)
def view_customer(request, customer_id):
    try:
        customer = Customer.objects.get(id=customer_id)
        return render(request, "admin/sales/customer/view_customer.html", {
            "customer": customer
        })
    except Customer.DoesNotExist:
        messages.error(request, "Customer not found.")
        return redirect("sales:customer_list")

def edit_customer(request, customer_id):
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        messages.error(request, "Customer not found.")
        return redirect("sales:customer_list")
    
    if request.method == "POST":
        try:
            same_as_whatsapp = request.POST.get("same_as_whatsapp") == "on"
            contact_number = request.POST.get("contact_number", "").strip()
            
            # Validate required fields
            first_name = request.POST.get("first_name", "").strip()
            display_name = request.POST.get("display_name", "").strip()
            customer_type = request.POST.get("customer_type", "Individual")
            
            if not first_name or not display_name:
                return render(request, "admin/sales/customer/edit_customer.html", {
                    "customer": customer,
                    "error": "First Name and Display Name are required fields."
                })
            
            if same_as_whatsapp:
                whatsapp_number = contact_number
            else:
                whatsapp_number = request.POST.get("whatsapp_number", "").strip()
            
            # Update customer
            customer.customer_type = customer_type
            customer.salutation = request.POST.get("salutation", "")
            customer.first_name = first_name
            customer.last_name = request.POST.get("last_name", "").strip()
            customer.display_name = display_name
            customer.place = request.POST.get("place", "").strip()
            customer.contact_number = contact_number
            customer.same_as_whatsapp = same_as_whatsapp
            customer.whatsapp_number = whatsapp_number
            customer.work_number = request.POST.get("work_number", "").strip()
            customer.gst_number = request.POST.get("gst_number", "").strip()
            
            customer.save()
            
            messages.success(request, f"Customer '{display_name}' has been updated successfully!")
            return redirect("sales:customer_list")
            
        except Exception as e:
            return render(request, "admin/sales/customer/edit_customer.html", {
                "customer": customer,
                "error": f"Error updating customer: {str(e)}"
            })
    
    return render(request, "admin/sales/customer/edit_customer.html", {
        "customer": customer
    })

def delete_customer(request, customer_id):
    if request.method != 'POST':
        return redirect('sales:customer_list')
    
    try:
        customer = Customer.objects.get(id=customer_id)
        customer_name = customer.display_name
        customer.delete()
        messages.success(request, f"Customer '{customer_name}' has been deleted successfully!")
    except Customer.DoesNotExist:
        messages.error(request, "Customer not found.")
    except Exception as e:
        messages.error(request, f"Error deleting customer: {str(e)}")
    
    return redirect("sales:customer_list")


# RESORT VIEWS
def resort_list(request):
    search_query = request.GET.get('search', '').strip()
    resorts = Resort.objects.all()
    
    if search_query:
        resorts = resorts.filter(
            Q(resort_name__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(contact_person__icontains=search_query)
        )
    
    return render(request, "admin/sales/resort/resort.html", {"resorts": resorts, "search_query": search_query})

def add_resort(request):
    if request.method == "POST":
        try:
            resort_name = request.POST.get("resort_name", "").strip()
            location = request.POST.get("location", "").strip()
            contact_person = request.POST.get("contact_person", "").strip()
            contact_number = request.POST.get("contact_number", "").strip()
            email = request.POST.get("email", "").strip()
            address = request.POST.get("address", "").strip()
            status = request.POST.get("status", "Active")
            
            if not resort_name or not location:
                messages.error(request, "Resort Name and Location are required.")
                return render(request, "admin/sales/resort/add_resort.html")
            
            if Resort.objects.filter(resort_name=resort_name).exists():
                messages.error(request, "Resort name already exists.")
                return render(request, "admin/sales/resort/add_resort.html")
            
            Resort.objects.create(
                resort_name=resort_name,
                location=location,
                contact_person=contact_person,
                contact_number=contact_number,
                email=email,
                address=address,
                status=status
            )
            messages.success(request, f"Resort '{resort_name}' added successfully!")
            return redirect("sales:resort_list")
        except Exception as e:
            messages.error(request, f"Error adding resort: {str(e)}")
            return render(request, "admin/sales/resort/add_resort.html")
    return render(request, "admin/sales/resort/add_resort.html")

def view_resort(request, resort_id):
    try:
        resort = Resort.objects.get(id=resort_id)
        return render(request, "admin/sales/resort/view_resort.html", {"resort": resort})
    except Resort.DoesNotExist:
        messages.error(request, "Resort not found.")
        return redirect("sales:resort_list")

def edit_resort(request, resort_id):
    try:
        resort = Resort.objects.get(id=resort_id)
    except Resort.DoesNotExist:
        messages.error(request, "Resort not found.")
        return redirect("sales:resort_list")
    
    if request.method == "POST":
        try:
            resort.resort_name = request.POST.get("resort_name", "").strip()
            resort.location = request.POST.get("resort_place", "").strip()  # Changed from "location" to "resort_place"
            resort.contact_person = request.POST.get("contact_person", "").strip()
            resort.contact_number = request.POST.get("contact_number", "").strip()
            resort.email = request.POST.get("email", "").strip()
            resort.address = request.POST.get("address", "").strip()
            resort.status = request.POST.get("status", "Active")
            resort.save()
            messages.success(request, f"Resort updated successfully!")
            return redirect("sales:resort_list")
        except Exception as e:
            return render(request, "admin/sales/resort/edit_resort.html", {
                "resort": resort,
                "error": f"Error updating resort: {str(e)}"
            })
    return render(request, "admin/sales/resort/edit_resort.html", {"resort": resort})

def delete_resort(request, resort_id):
    if request.method != 'POST':
        return redirect('sales:resort_list')
    
    try:
        resort = Resort.objects.get(id=resort_id)
        resort_name = resort.resort_name
        resort.delete()
        messages.success(request, f"Resort '{resort_name}' deleted successfully!")
    except Resort.DoesNotExist:
        messages.error(request, "Resort not found.")
    except Exception as e:
        messages.error(request, f"Error deleting resort: {str(e)}")
    return redirect("sales:resort_list")


# MEAL VIEWS
def meal_list(request):
    search_query = request.GET.get('search', '').strip()
    meals = Meal.objects.all()
    
    if search_query:
        meals = meals.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    return render(request, "admin/sales/meals/meals.html", {"meals": meals, "search_query": search_query})

def add_meal(request):
    if request.method == "POST":
        try:
            name = request.POST.get("name", "").strip()
            description = request.POST.get("description", "").strip()
            included_meals = request.POST.getlist("included_meals")
            included_meals_str = ", ".join(included_meals) if included_meals else ""
            
            if not name:
                messages.error(request, "Meal Plan Name is required.")
                return render(request, "admin/sales/meals/add_meals.html")
            
            # Validate name length
            if len(name) < 2 or len(name) > 50:
                messages.error(request, "Meal plan name must be between 2 and 50 characters.")
                return render(request, "admin/sales/meals/add_meals.html")
            
            # Check if at least one meal is selected
            if not included_meals:
                messages.error(request, "Please select at least one meal type.")
                return render(request, "admin/sales/meals/add_meals.html")
            
            if Meal.objects.filter(name=name).exists():
                messages.error(request, "Meal plan name already exists.")
                return render(request, "admin/sales/meals/add_meals.html")
            
            Meal.objects.create(
                name=name,
                description=description,
                included_meals=included_meals_str
            )
            messages.success(request, f"Meal plan '{name}' added successfully!")
            return redirect("sales:meal_list")
        except Exception as e:
            messages.error(request, f"Error adding meal plan: {str(e)}")
            return render(request, "admin/sales/meals/add_meals.html")
    return render(request, "admin/sales/meals/add_meals.html")

def view_meal(request, meal_id):
    try:
        meal = Meal.objects.get(id=meal_id)
        return render(request, "admin/sales/meals/view_meals.html", {"meal": meal})
    except Meal.DoesNotExist:
        messages.error(request, "Meal not found.")
        return redirect("sales:meal_list")

def edit_meal(request, meal_id):
    try:
        meal = Meal.objects.get(id=meal_id)
    except Meal.DoesNotExist:
        messages.error(request, "Meal not found.")
        return redirect("sales:meal_list")
    
    if request.method == "POST":
        try:
            meal.name = request.POST.get("name", "").strip()
            meal.description = request.POST.get("description", "").strip()
            included_meals = request.POST.getlist("included_meals")
            meal.included_meals = ", ".join(included_meals) if included_meals else ""
            meal.status = request.POST.get("status", "Available")
            meal.save()
            messages.success(request, f"Meal '{meal.name}' updated successfully!")
            return redirect("sales:meal_list")
        except Exception as e:
            return render(request, "admin/sales/meals/edit_meals.html", {
                "meal": meal,
                "error": f"Error updating meal: {str(e)}"
            })
    return render(request, "admin/sales/meals/edit_meals.html", {"meal": meal})

def delete_meal(request, meal_id):
    if request.method != 'POST':
        return redirect('sales:meal_list')
    
    try:
        meal = Meal.objects.get(id=meal_id)
        meal_name = meal.name
        meal.delete()
        messages.success(request, f"Meal '{meal_name}' deleted successfully!")
    except Meal.DoesNotExist:
        messages.error(request, "Meal not found.")
    except Exception as e:
        messages.error(request, f"Error deleting meal: {str(e)}")
    return redirect("sales:meal_list")


# VOUCHER VIEWS
def voucher_list(request):
    search_query = request.GET.get('search', '').strip()
    vouchers = Voucher.objects.all()
    
    if search_query:
        vouchers = vouchers.filter(
            Q(voucher_no__icontains=search_query) |
            Q(customer_display_name_icontains=search_query) |
            Q(resort_resort_name_icontains=search_query)
        )
    
    vouchers = vouchers.order_by('-voucher_date', '-id')
    return render(request, "admin/sales/vouchers/vouchers.html", {"vouchers": vouchers, "search_query": search_query})

def add_voucher(request):
    customers = Customer.objects.all()
    resorts = Resort.objects.all()
    accounts = Account.objects.all()
    employees = Employee.objects.filter(status='Active')
    meals = Meal.objects.all()
    
    if request.method == "POST":
        try:
            cus_id = request.POST.get("customer_id")
            vno = request.POST.get("voucher_no", "").strip()
            vdate = request.POST.get("voucher_date")

            # required validation
            if not cus_id or not vno or not vdate:
                messages.error(request, "Customer, voucher number, and date are required.")
                return render(request, "admin/sales/vouchers/add_vouchers.html", {"customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees, "meals": meals})

            if Voucher.objects.filter(voucher_no=vno).exists():
                messages.error(request, "Voucher number already exists.")
                return render(request, "admin/sales/vouchers/add_vouchers.html", {"customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees, "meals": meals})

            Voucher.objects.create(
                customer_id=cus_id,
                voucher_no=vno,
                voucher_date=vdate,
                sales_person_id=request.POST.get("sales_person") or None,
                resort_id=request.POST.get("resort") or None,
                checkin_date=request.POST.get("checkin_date"),
                checkout_date=request.POST.get("checkout_date"),
                checkin_time=request.POST.get("checkin_time"),
                checkout_time=request.POST.get("checkout_time"),
                adults=request.POST.get("adults", 0),
                children=request.POST.get("children", 0),
                nights=request.POST.get("nights", 1),
                pax_notes=request.POST.get("pax_notes", "").strip(),
                room_type=request.POST.get("room_type", "").strip(),
                no_of_rooms=request.POST.get("no_of_rooms", 1),
                meals_plan_id=request.POST.get("meals_plan") or None,
                bank_account_id=request.POST.get("bank_account") or None,
                package_price=request.POST.get("package_price", 0),
                resort_price=request.POST.get("resort_price", 0),
                total_amount=request.POST.get("total_amount", 0),
                received=request.POST.get("received", 0),
                pending=request.POST.get("pending", 0),
                from_whytehouse=request.POST.get("from_whytehouse", 0),
                profit=request.POST.get("profit", 0),
                note_for_resort=request.POST.get("note_for_resort", "").strip(),
                note_for_guest=request.POST.get("note_for_guest", "").strip()
            )
            messages.success(request, f"Voucher added successfully!")
            return redirect("sales:voucher_list")
        except Exception as e:
            messages.error(request, f"Error adding voucher: {str(e)}")
    return render(request, "admin/sales/vouchers/add_vouchers.html", {"customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees, "meals": meals})

def view_voucher(request, voucher_id):
    try:
        voucher = Voucher.objects.get(id=voucher_id)
        return render(request, "admin/sales/vouchers/view_vouchers.html", {"voucher": voucher})
    except Voucher.DoesNotExist:
        messages.error(request, "Voucher not found.")
        return redirect("sales:voucher_list")

def edit_voucher(request, voucher_id):
    try:
        voucher = Voucher.objects.get(id=voucher_id)
    except Voucher.DoesNotExist:
        messages.error(request, "Voucher not found.")
        return redirect("sales:voucher_list")
    
    if request.method == "POST":
        try:
            # don't allow editing voucher_no or voucher_date here; keep current values
            voucher.voucher_code = request.POST.get("voucher_code", "").strip()
            voucher.discount_amount = request.POST.get("discount_amount", 0)
            voucher.discount_percentage = request.POST.get("discount_percentage", 0) or None
            voucher.description = request.POST.get("description", "").strip()
            voucher.valid_from = request.POST.get("valid_from")
            voucher.valid_till = request.POST.get("valid_till")
            voucher.status = request.POST.get("status", "Active")
            voucher.save()
            messages.success(request, f"Voucher updated successfully!")
            return redirect("sales:voucher_list")
        except Exception as e:
            return render(request, "admin/sales/vouchers/edit_vouchers.html", {
                "voucher": voucher,
                "error": f"Error updating voucher: {str(e)}"
            })
    return render(request, "admin/sales/vouchers/edit_vouchers.html", {"voucher": voucher})

def delete_voucher(request, voucher_id):
    if request.method != 'POST':
        return redirect('sales:voucher_list')
    
    try:
        voucher = Voucher.objects.get(id=voucher_id)
        voucher_no = voucher.voucher_no
        voucher.delete()
        messages.success(request, f"Voucher '{voucher_no}' deleted successfully!")
    except Voucher.DoesNotExist:
        messages.error(request, "Voucher not found.")
    except Exception as e:
        messages.error(request, f"Error deleting voucher: {str(e)}")
    return redirect("sales:voucher_list")


# INVOICE VIEWS
def invoice_list(request):
    search_query = request.GET.get('search', '').strip()
    invoices = Invoice.objects.all()
    
    if search_query:
        invoices = invoices.filter(
            Q(invoice_no__icontains=search_query) |
            Q(customer_display_name_icontains=search_query) |
            Q(resort_resort_name_icontains=search_query)
        )
    
    invoices = invoices.order_by('-invoice_date', '-id')
    return render(request, "admin/sales/invoice/invoice.html", {"invoices": invoices, "search_query": search_query})

def add_invoice(request):
    customers = Customer.objects.all()
    resorts = Resort.objects.all()
    accounts = Account.objects.all()
    employees = Employee.objects.filter(status='Active')
    meals = Meal.objects.all()
    
    if request.method == "POST":
        try:
            customer_id = request.POST.get("customer_id")
            invoice_no = request.POST.get("invoice_no", "").strip()
            invoice_date = request.POST.get("invoice_date")
            sales_person_id = request.POST.get("sales_person")
            resort_id = request.POST.get("resort")
            checkin_date = request.POST.get("checkin_date") or None
            checkout_date = request.POST.get("checkout_date") or None
            checkin_time = request.POST.get("checkin_time") or None
            checkout_time = request.POST.get("checkout_time") or None
            adults = request.POST.get("adults", 0)
            children = request.POST.get("children", 0)
            pax_total = request.POST.get("pax_total", 0)
            pax_notes = request.POST.get("pax_notes", "").strip()
            nights = request.POST.get("nights", 1)
            room_type = request.POST.get("room_type", "").strip()
            rooms = request.POST.get("rooms", 1)
            meals_plan = request.POST.get("meals_plan", "").strip()
            bank_account_id = request.POST.get("bank_account")
            package_price = request.POST.get("package_price", 0)
            tax = request.POST.get("tax", 0)
            resort_price = request.POST.get("resort_price", 0)
            total = request.POST.get("total", 0)
            received = request.POST.get("received", 0)
            pending = request.POST.get("pending", 0)
            profit = request.POST.get("profit", 0)
            notes = request.POST.get("notes", "").strip()
            
            if not customer_id or not invoice_no or not invoice_date:
                messages.error(request, "Customer, Invoice No, and Invoice Date are required.")
                return render(request, "admin/sales/invoice/add_invoice.html", {"customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees})
            
            if Invoice.objects.filter(invoice_no=invoice_no).exists():
                messages.error(request, "Invoice number already exists.")
                return render(request, "admin/sales/invoice/add_invoice.html", {"customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees})
            
            Invoice.objects.create(
                customer_id=customer_id,
                invoice_no=invoice_no,
                invoice_date=invoice_date,
                sales_person_id=sales_person_id if sales_person_id else None,
                resort_id=resort_id if resort_id else None,
                checkin_date=checkin_date,
                checkout_date=checkout_date,
                checkin_time=checkin_time,
                checkout_time=checkout_time,
                adults=adults,
                children=children,
                pax_total=pax_total,
                pax_notes=pax_notes,
                nights=nights,
                room_type=room_type,
                rooms=rooms,
                meals_plan=meals_plan,
                bank_account_id=bank_account_id if bank_account_id else None,
                package_price=package_price,
                tax=tax,
                resort_price=resort_price,
                total=total,
                received=received,
                pending=pending,
                profit=profit,
                notes=notes
            )
            messages.success(request, f"Invoice '{invoice_no}' added successfully!")
            return redirect("sales:invoice_list")
        except Exception as e:
            messages.error(request, f"Error adding invoice: {str(e)}")
            return render(request, "admin/sales/invoice/add_invoice.html", {"customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees, "meals": meals})
    return render(request, "admin/sales/invoice/add_invoice.html", {"customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees, "meals": meals})

def view_invoice(request, invoice_id):
    try:
        invoice = Invoice.objects.get(id=invoice_id)
        return render(request, "admin/sales/invoice/view_invoice.html", {"invoice": invoice})
    except Invoice.DoesNotExist:
        messages.error(request, "Invoice not found.")
        return redirect("sales:invoice_list")

def edit_invoice(request, invoice_id):
    customers = Customer.objects.all()
    resorts = Resort.objects.all()
    accounts = Account.objects.all()
    employees = Employee.objects.filter(status='Active')
    
    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        messages.error(request, "Invoice not found.")
        return redirect("sales:invoice_list")
    
    if request.method == "POST":
        try:
            cust = request.POST.get("customer_id")
            ino = request.POST.get("invoice_no", "").strip()
            idate = request.POST.get("invoice_date")
            if not cust or not ino or not idate:
                messages.error(request, "Customer, Invoice No, and Invoice Date are required.")
                return render(request, "admin/sales/invoice/edit_invoice.html", {"invoice": invoice, "customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees})

            invoice.customer_id = cust
            invoice.invoice_no = ino
            invoice.invoice_date = idate
            invoice.sales_person_id = request.POST.get("sales_person") or None
            invoice.resort_id = request.POST.get("resort") or None
            invoice.checkin_date = request.POST.get("checkin_date") or None
            invoice.checkout_date = request.POST.get("checkout_date") or None
            invoice.checkin_time = request.POST.get("checkin_time") or None
            invoice.checkout_time = request.POST.get("checkout_time") or None
            invoice.adults = request.POST.get("adults", 0)
            invoice.children = request.POST.get("children", 0)
            invoice.pax_total = request.POST.get("pax_total", 0)
            invoice.pax_notes = request.POST.get("pax_notes", "").strip()
            invoice.nights = request.POST.get("nights", 1)
            invoice.room_type = request.POST.get("room_type", "").strip()
            invoice.rooms = request.POST.get("rooms", 1)
            invoice.meals_plan = request.POST.get("meals_plan", "").strip()
            invoice.bank_account_id = request.POST.get("bank_account") or None
            invoice.package_price = request.POST.get("package_price", 0)
            invoice.tax = request.POST.get("tax", 0)
            invoice.resort_price = request.POST.get("resort_price", 0)
            invoice.total = request.POST.get("total", 0)
            invoice.received = request.POST.get("received", 0)
            invoice.pending = request.POST.get("pending", 0)
            invoice.profit = request.POST.get("profit", 0)
            invoice.notes = request.POST.get("notes", "").strip()
            invoice.save()
            
            messages.success(request, f"Invoice '{invoice.invoice_no}' updated successfully!")
            return redirect("sales:invoice_list")
        except Exception as e:
            return render(request, "admin/sales/invoice/edit_invoice.html", {
                "invoice": invoice,
                "customers": customers,
                "resorts": resorts,
                "accounts": accounts,
                "employees": employees,
                "error": f"Error updating invoice: {str(e)}"
            })
    return render(request, "admin/sales/invoice/edit_invoice.html", {"invoice": invoice, "customers": customers, "resorts": resorts, "accounts": accounts, "employees": employees})

def delete_invoice(request, invoice_id):
    if request.method != 'POST':
        return redirect('sales:invoice_list')
    
    try:
        invoice = Invoice.objects.get(id=invoice_id)
        invoice_no = invoice.invoice_no
        invoice.delete()
        messages.success(request, f"Invoice '{invoice_no}' deleted successfully!")
    except Invoice.DoesNotExist:
        messages.error(request, "Invoice not found.")
    except Exception as e:
        messages.error(request, f"Error deleting invoice: {str(e)}")
    return redirect("sales:invoice_list")

def blog_list(request):
    blogs = Blog.objects.all()

    search_query = request.GET.get("search", "").strip()
    if search_query:
        blogs = blogs.filter(
            Q(title__icontains=search_query) |
            Q(author_name__icontains=search_query) |
            Q(slug__icontains=search_query) |
            Q(status__icontains=search_query)
        )

    status_filter = request.GET.get("status", "").strip()
    if status_filter:
        blogs = blogs.filter(status=status_filter)

    published_count = blogs.filter(status="published").count()
    draft_count = blogs.filter(status="draft").count()

    context = {
        "blogs": blogs,
        "search_query": search_query,
        "status_filter": status_filter,
        "published_count": published_count,
        "draft_count": draft_count,
        "now": datetime.now().strftime("%B %d, %Y"),
    }
    return render(request, "admin/blog/blog.html", context)


def add_blog(request):
    categories = BlogCategory.objects.all().order_by("name")

    if request.method == "POST":
        try:
            title = (request.POST.get("title") or "").strip()
            slug = (request.POST.get("slug") or "").strip()
            excerpt = (request.POST.get("excerpt") or "").strip()
            content = (request.POST.get("content") or "").strip()
            author_name = (request.POST.get("author_name") or "").strip()
            author_summary = (request.POST.get("author_summary") or "").strip()
            reading_time = request.POST.get("reading_time")
            publish_date = request.POST.get("publish_date")

            if not title or not slug or not excerpt or not content or not author_name or not author_summary or not reading_time or not publish_date:
                messages.error(request, "Please fill in all required fields.")
                return render(request, "admin/blog/add_blog.html", {"categories": categories})

            import re
            if not re.match(r'^[a-z0-9-]+$', slug):
                messages.error(request, "Slug must contain only lowercase letters, numbers, and hyphens.")
                return render(request, "admin/blog/add_blog.html", {"categories": categories})

            if Blog.objects.filter(slug=slug).exists():
                messages.error(request, "A blog with this slug already exists.")
                return render(request, "admin/blog/add_blog.html", {"categories": categories})

            try:
                reading_time_val = int(reading_time)
                if reading_time_val < 1 or reading_time_val > 120:
                    messages.error(request, "Reading time must be between 1 and 120 minutes.")
                    return render(request, "admin/blog/add_blog.html", {"categories": categories})
            except ValueError:
                messages.error(request, "Invalid reading time format.")
                return render(request, "admin/blog/add_blog.html", {"categories": categories})

            if len(excerpt) < 50 or len(excerpt) > 300:
                messages.error(request, "Excerpt must be between 50 and 300 characters.")
                return render(request, "admin/blog/add_blog.html", {"categories": categories})

            hashtags_value = (request.POST.get("hashtags") or "").strip()
            category_slug = request.POST.get("category", "")

            blog = Blog.objects.create(
                title=title,
                slug=slug,
                excerpt=excerpt,
                content=content,
                status=request.POST.get("status", "draft"),
                category=category_slug,
                package_id=(request.POST.get("package_id") or "").strip() or None,
                author_name=author_name,
                author_summary=author_summary,
                reading_time=reading_time_val,
                publish_date=publish_date,
                tags=hashtags_value,
            )

            if request.FILES.get("featured_image"):
                blog.featured_image = request.FILES["featured_image"]
                blog.save()

            content_images = request.FILES.getlist("content_images")
            for idx, image_file in enumerate(content_images):
                BlogImage.objects.create(blog=blog, image=image_file, order=idx)

            messages.success(request, "Blog created successfully!")
            return redirect("blog:blog_list")

        except Exception as e:
            messages.error(request, f"Error creating blog: {str(e)}")
            return render(request, "admin/blog/add_blog.html", {"categories": categories})

    return render(request, "admin/blog/add_blog.html", {"categories": categories})

def edit_blog(request, blog_id):
    blog = get_object_or_404(Blog, id=blog_id)
    categories = BlogCategory.objects.filter(is_active=True).order_by('order', 'name')

    if request.method == "POST":
        try:
            title = (request.POST.get("title") or "").strip()
            slug = (request.POST.get("slug") or "").strip()
            excerpt = (request.POST.get("excerpt") or "").strip()
            content = (request.POST.get("content") or "").strip()
            author_name = (request.POST.get("author_name") or "").strip()
            author_summary = (request.POST.get("author_summary") or "").strip()
            reading_time = request.POST.get("reading_time")
            publish_date = request.POST.get("publish_date")

            if not title or not slug or not excerpt or not content or not author_name or not author_summary or not reading_time or not publish_date:
                messages.error(request, "Please fill in all required fields.")
                return render(request, "admin/blog/edit_blog.html", {"blog": blog, "categories": categories})

            hashtags_value = (request.POST.get("hashtags") or "").strip()
            
            # Get category slug from dropdown
            category_slug = request.POST.get("category", "")
            
            blog.title = title
            blog.slug = slug
            blog.excerpt = excerpt
            blog.content = content
            blog.status = request.POST.get("status", "draft")
            blog.category = category_slug
            blog.package_id = (request.POST.get("package_id") or "").strip() or None

            blog.author_name = author_name
            blog.author_summary = author_summary
            blog.reading_time = int(reading_time or 1)
            blog.publish_date = publish_date

            blog.featured_image_url = (request.POST.get("featured_image_url") or "").strip() or None
            blog.hashtags = hashtags_value
            blog.tags = hashtags_value  # Save to both fields for compatibility

            if request.FILES.get("featured_image"):
                blog.featured_image = request.FILES["featured_image"]

            blog.save()

            # Handle deleted content images
            deleted_images = request.POST.get("deleted_images", "")
            if deleted_images:
                deleted_ids = [int(id) for id in deleted_images.split(",") if id.strip()]
                BlogImage.objects.filter(id__in=deleted_ids, blog=blog).delete()
                
                # Reorder remaining images
                remaining_images = blog.images.all().order_by('order')
                for idx, img in enumerate(remaining_images):
                    img.order = idx
                    img.save()

            # Handle new content images
            content_images = request.FILES.getlist("content_images")
            if content_images:
                # Get current max order
                current_max = blog.images.count()
                for idx, image_file in enumerate(content_images):
                    BlogImage.objects.create(
                        blog=blog,
                        image=image_file,
                        order=current_max + idx
                    )

            messages.success(request, "Blog updated successfully!")
            return redirect("blog:blog_list")

        except Exception as e:
            messages.error(request, f"Error updating blog: {str(e)}")

    return render(request, "admin/blog/edit_blog.html", {"blog": blog, "categories": categories})


def view_blog(request, slug):
    blog = get_object_or_404(Blog, slug=slug)

    tags = []
    if blog.hashtags:
        tags = [t.strip() for t in blog.hashtags.split(",") if t.strip()]

    # Replace {{image1}}, {{image2}}, etc. with actual image HTML
    content = blog.content
    content_images = blog.images.all()
    
    # Normalize line endings and split content
    content = content.replace('\r\n', '\n').replace('\r', '\n')
    
    # Split by lines and process
    lines = content.split('\n')
    processed_lines = []
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        
        if not line:
            i += 1
            continue
        
        # Check if this line contains an image tag
        has_image = False
        
        for idx, img in enumerate(content_images):
            image_num = idx + 1
            
            # Check for left positioned image
            placeholder_left = f"{{{{image{image_num}-left}}}}"
            if placeholder_left in line:
                # Get the text before the image tag (could be on same line or previous lines)
                text_content = line.replace(placeholder_left, '').strip()
                
                # If no text on this line, look at previous non-empty lines
                if not text_content and processed_lines:
                    # Get previous paragraph
                    prev_content = []
                    while processed_lines and not processed_lines[-1].startswith('<'):
                        prev_content.insert(0, processed_lines.pop())
                    text_content = ' '.join(prev_content)
                
                # Create side-by-side layout with image on left
                processed_lines.append(f'''
                    <div class="content-row">
                        <figure class="content-image-left">
                            <img src="{img.image.url}" alt="Content image {image_num}">
                        </figure>
                        <div class="content-text"><p>{text_content}</p></div>
                    </div>
                ''')
                has_image = True
                break
            
            # Check for right positioned image
            placeholder_right = f"{{{{image{image_num}-right}}}}"
            if placeholder_right in line:
                # Get the text before the image tag
                text_content = line.replace(placeholder_right, '').strip()
                
                # If no text on this line, look at previous non-empty lines
                if not text_content and processed_lines:
                    # Get previous paragraph
                    prev_content = []
                    while processed_lines and not processed_lines[-1].startswith('<'):
                        prev_content.insert(0, processed_lines.pop())
                    text_content = ' '.join(prev_content)
                
                # Create side-by-side layout with image on right
                processed_lines.append(f'''
                    <div class="content-row">
                        <div class="content-text"><p>{text_content}</p></div>
                        <figure class="content-image-right">
                            <img src="{img.image.url}" alt="Content image {image_num}">
                        </figure>
                    </div>
                ''')
                has_image = True
                break
            
            # Check for center positioned image
            placeholder_center = f"{{{{image{image_num}-center}}}}"
            if placeholder_center in line:
                text_content = line.replace(placeholder_center, '').strip()
                if text_content:
                    processed_lines.append(f'<p>{text_content}</p>')
                processed_lines.append(f'''
                    <figure class="content-image-center">
                        <img src="{img.image.url}" alt="Content image {image_num}">
                    </figure>
                ''')
                has_image = True
                break
            
            # Check for default (center) positioned image
            placeholder_default = f"{{{{image{image_num}}}}}"
            if placeholder_default in line:
                text_content = line.replace(placeholder_default, '').strip()
                if text_content:
                    processed_lines.append(f'<p>{text_content}</p>')
                processed_lines.append(f'''
                    <figure class="content-image-center">
                        <img src="{img.image.url}" alt="Content image {image_num}">
                    </figure>
                ''')
                has_image = True
                break
        
        # If no image found in this line, add it as regular text
        if not has_image and line:
            processed_lines.append(line)
        
        i += 1
    
    content = '\n'.join(processed_lines)

    return render(request, "admin/blog/view_blog.html", {"blog": blog, "tags": tags, "processed_content": content})


def delete_blog(request, blog_id):
    if request.method != "POST":
        return redirect("blog:blog_list")

    blog = get_object_or_404(Blog, id=blog_id)
    title = blog.title
    blog.delete()
    messages.success(request, f"Blog '{title}' deleted successfully!")
    return redirect("blog:blog_list")

def toggle_blog_status(request, blog_id):
    if request.method != 'POST':
        return redirect('blog:blog_list')
    
    blog = get_object_or_404(Blog, id=blog_id)
    
    if blog.status == 'published':
        blog.status = 'draft'
        messages.success(request, f"Blog '{blog.title}' unpublished successfully!")
    else:
        blog.status = 'published'
        messages.success(request, f"Blog '{blog.title}' published successfully!")
    
    blog.save()
    return redirect('blog:blog_list')


def add_category(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        if not name:
            messages.error(request, "Category name is required.")
            return redirect("blog:add_category")

        slug = slugify(name)

        # ensure unique slug
        base_slug = slug
        i = 1
        while BlogCategory.objects.filter(slug=slug).exists():
            slug = f"{base_slug}-{i}"
            i += 1

        BlogCategory.objects.create(name=name, slug=slug)
        messages.success(request, f"Category '{name}' added successfully!")
        
        return redirect("blog:add_category")

    categories = BlogCategory.objects.all().order_by("name")
    return render(request, "admin/blog/manage_categories.html", {"categories": categories})
def delete_category(request, category_id):
    try:
        category = get_object_or_404(BlogCategory, id=category_id)
        category_name = category.name
        category.delete()
        messages.success(request, f"Category '{category_name}' deleted successfully.")
    except:
        messages.warning(request, "Category not found or already deleted.")
    return redirect("blog:add_category")

def customer_report(request):
    customers = Customer.objects.all().order_by("-created_at")

    if request.method == "POST" and request.POST.get("action") == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"

        headers = ["Sl No", "First Name", "Last Name", "Display Name", "Place", "Mobile", "WhatsApp"]
        ws.append(headers)

        for i, c in enumerate(customers, start=1):
            ws.append([
                i,
                c.first_name or "",
                c.last_name or "",
                c.display_name or "",
                c.place or "",
                c.contact_number or "",
                c.whatsapp_number or "",
            ])

        # Auto width
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 3, 45)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"customer_report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    return render(request, "admin/report/customer_report.html", {"customers": customers})

def invoice_report(request):
    resorts = Resort.objects.all().order_by("resort_name")
    invoices = Invoice.objects.none()
    employee_view = False
    employees = Employee.objects.none()

    selected = {"from_date": "", "to_date": "", "resort": "", "employee": ""}

    if request.method == "POST":
        action = request.POST.get("action")
        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")
        resort_id = request.POST.get("resort")
        employee_view = request.POST.get("employee_view") == "on"
        employee_id = request.POST.get("employee")

        selected = {
            "from_date": from_date or "",
            "to_date": to_date or "",
            "resort": resort_id or "",
            "employee": employee_id or "",
        }

        if from_date and to_date and resort_id:
            try:
                fd = datetime.strptime(from_date, "%Y-%m-%d").date()
                td = datetime.strptime(to_date, "%Y-%m-%d").date()
                
                # Filter employees who have invoices in the selected date range and resort
                if employee_view:
                    employee_ids = Invoice.objects.filter(
                        resort_id=resort_id,
                        invoice_date__range=(fd, td),
                        sales_person__isnull=False
                    ).values_list('sales_person_id', flat=True).distinct()
                    
                    employees = Employee.objects.filter(
                        id__in=employee_ids,
                        status="Active"
                    ).order_by("name")

                # Base queryset
                qs = Invoice.objects.select_related(
                    "customer", "resort", "sales_person", "bank_account"
                ).filter(resort_id=resort_id, invoice_date__range=(fd, td))

                # Employee filter
                if employee_view and employee_id:
                    qs = qs.filter(sales_person_id=employee_id)

                invoices = qs.order_by("-invoice_date", "-id")

                # Excel export
                if action == "excel":
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Invoice Report"

                    headers = [
                        "Invoice No", "Invoice Date",
                        "Customer", "Mobile",
                        "Sales Person",
                        "Resort",
                        "Check-in Date", "Check-out Date",
                        "Check-in Time", "Check-out Time",
                        "Adults", "Children", "Pax Total", "Pax Notes",
                        "Nights", "Room Type", "Rooms", "Meals Plan",
                        "Bank Account",
                        "Package Price", "Tax", "Resort Price",
                        "Total", "Received", "Pending", "Profit",
                        "Notes",
                    ]

                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for row, inv in enumerate(invoices, start=2):
                        ws.cell(row=row, column=1, value=inv.invoice_no)
                        ws.cell(row=row, column=2, value=inv.invoice_date.strftime("%d/%m/%Y") if inv.invoice_date else "")

                        ws.cell(row=row, column=3, value=inv.customer.display_name if inv.customer else "")
                        ws.cell(row=row, column=4, value=getattr(inv.customer, "contact_number", "") if inv.customer else "")

                        ws.cell(row=row, column=5, value=inv.sales_person.name if inv.sales_person else "")
                        ws.cell(row=row, column=6, value=inv.resort.resort_name if inv.resort else "")

                        ws.cell(row=row, column=7, value=inv.checkin_date.strftime("%d/%m/%Y") if inv.checkin_date else "")
                        ws.cell(row=row, column=8, value=inv.checkout_date.strftime("%d/%m/%Y") if inv.checkout_date else "")

                        ws.cell(row=row, column=9, value=str(inv.checkin_time) if inv.checkin_time else "")
                        ws.cell(row=row, column=10, value=str(inv.checkout_time) if inv.checkout_time else "")

                        ws.cell(row=row, column=11, value=int(inv.adults or 0))
                        ws.cell(row=row, column=12, value=int(inv.children or 0))
                        ws.cell(row=row, column=13, value=int(inv.pax_total or 0))
                        ws.cell(row=row, column=14, value=inv.pax_notes or "")

                        ws.cell(row=row, column=15, value=int(inv.nights or 0))
                        ws.cell(row=row, column=16, value=inv.room_type or "")
                        ws.cell(row=row, column=17, value=int(inv.rooms or 0))
                        ws.cell(row=row, column=18, value=inv.meals_plan or "")

                        ws.cell(row=row, column=19, value=inv.bank_account.account_name if inv.bank_account else "")

                        ws.cell(row=row, column=20, value=float(inv.package_price or 0))
                        ws.cell(row=row, column=21, value=float(inv.tax or 0))
                        ws.cell(row=row, column=22, value=float(inv.resort_price or 0))

                        ws.cell(row=row, column=23, value=float(inv.total or 0))
                        ws.cell(row=row, column=24, value=float(inv.received or 0))
                        ws.cell(row=row, column=25, value=float(inv.pending or 0))
                        ws.cell(row=row, column=26, value=float(inv.profit or 0))

                        ws.cell(row=row, column=27, value=inv.notes or "")

                    # Auto width
                    for col in range(1, len(headers) + 1):
                        letter = get_column_letter(col)
                        max_len = 0
                        for c in ws[letter]:
                            if c.value:
                                max_len = max(max_len, len(str(c.value)))
                        ws.column_dimensions[letter].width = min(max_len + 3, 45)

                    response = HttpResponse(
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    filename = f"invoice_report_{fd.strftime('%d%m%Y')}_{td.strftime('%d%m%Y')}.xlsx"
                    response["Content-Disposition"] = f'attachment; filename="{filename}"'
                    wb.save(response)
                    return response

            except ValueError:
                messages.error(request, "Invalid date format.")

    return render(request, "admin/report/invoice_report.html", {
        "employees": employees,
        "resorts": resorts,
        "invoices": invoices,
        "employee_view": employee_view,
        "selected": selected,
    })



def voucher_report(request):
    resorts = Resort.objects.all().order_by("resort_name")
    vouchers = Voucher.objects.none()
    employee_view = False
    employees = Employee.objects.none()
    
    selected = {"from_date": "", "to_date": "", "resort": "", "employee": ""}
    
    if request.method == "POST":
        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")
        resort_id = request.POST.get("resort")
        employee_view = request.POST.get("employee_view") == "on"
        employee_id = request.POST.get("employee")
        action = request.POST.get("action")
        
        selected = {
            "from_date": from_date or "",
            "to_date": to_date or "",
            "resort": resort_id or "",
            "employee": employee_id or "",
        }
        
        if from_date and to_date and resort_id:
            try:
                fd = datetime.strptime(from_date, "%Y-%m-%d").date()
                td = datetime.strptime(to_date, "%Y-%m-%d").date()
                
                # Filter employees who have vouchers in the selected date range and resort
                if employee_view:
                    employee_ids = Voucher.objects.filter(
                        resort_id=resort_id,
                        voucher_date__range=(fd, td),
                        sales_person__isnull=False
                    ).values_list('sales_person_id', flat=True).distinct()
                    
                    employees = Employee.objects.filter(
                        id__in=employee_ids,
                        status="Active"
                    ).order_by("name")
                
                # Base queryset
                qs = Voucher.objects.select_related(
                    "customer", "resort", "sales_person", "bank_account", "meals_plan"
                ).filter(resort_id=resort_id, voucher_date__range=(fd, td))
                
                # Employee filter
                if employee_view and employee_id:
                    qs = qs.filter(sales_person_id=employee_id)
                
                vouchers = qs.order_by("-voucher_date", "-id")
                
                # Excel export
                if action == "excel":
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Voucher Report"
                    
                    # Header styling
                    header_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    # Headers
                    headers = ["Voucher No", "Date", "Customer", "Resort", "Total Amount"]
                    if employee_view:
                        headers.insert(3, "Sales Person")
                    
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Data rows
                    for row_num, voucher in enumerate(vouchers, 2):
                        ws.cell(row=row_num, column=1, value=voucher.voucher_no)
                        ws.cell(row=row_num, column=2, value=voucher.voucher_date.strftime("%d/%m/%Y"))
                        ws.cell(row=row_num, column=3, value=voucher.customer.display_name if voucher.customer else "-")
                        
                        if employee_view:
                            ws.cell(row=row_num, column=4, value=voucher.sales_person.name if voucher.sales_person else "-")
                            ws.cell(row=row_num, column=5, value=voucher.resort.resort_name if voucher.resort else "-")
                            ws.cell(row=row_num, column=6, value=float(voucher.total))
                        else:
                            ws.cell(row=row_num, column=4, value=voucher.resort.resort_name if voucher.resort else "-")
                            ws.cell(row=row_num, column=5, value=float(voucher.total))
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    response = HttpResponse(
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    response["Content-Disposition"] = f'attachment; filename="voucher_report_{fd.strftime("%d%m%Y")}_{td.strftime("%d%m%Y")}.xlsx"'
                    wb.save(response)
                    return response
                
            except ValueError:
                messages.error(request, "Invalid date format.")
    
    return render(request, "admin/report/voucher_report.html", {
        "resorts": resorts,
        "vouchers": vouchers,
        "employees": employees,
        "employee_view": employee_view,
        "selected": selected,
    })

def leads_report(request):
    leads = Lead.objects.none()
    selected = {"from_date": "", "to_date": "", "enquiry_type": ""}
    
    if request.method == "POST":
        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")
        enquiry_type = request.POST.get("enquiry_type")
        action = request.POST.get("action")
        
        selected = {
            "from_date": from_date or "",
            "to_date": to_date or "",
            "enquiry_type": enquiry_type or "",
        }
        
        if from_date and to_date:
            try:
                fd = datetime.strptime(from_date, "%Y-%m-%d").date()
                td = datetime.strptime(to_date, "%Y-%m-%d").date()
                
                # Base queryset
                qs = Lead.objects.filter(created_at__date__range=(fd, td))
                
                # Enquiry type filter
                if enquiry_type:
                    qs = qs.filter(enquiry_type=enquiry_type)
                
                leads = qs.order_by("-created_at")
                
                # Excel export
                if action == "excel":
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Leads Report"
                    
                    # Header styling
                    header_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    # Headers
                    headers = ["Full Name", "Mobile Number", "Place", "Source", "Enquiry Type", "Remarks", "Created Date"]
                    
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Data rows
                    for row_num, lead in enumerate(leads, 2):
                        ws.cell(row=row_num, column=1, value=lead.full_name)
                        ws.cell(row=row_num, column=2, value=lead.mobile_number)
                        ws.cell(row=row_num, column=3, value=lead.place or "-")
                        ws.cell(row=row_num, column=4, value=lead.get_source_display())
                        ws.cell(row=row_num, column=5, value=lead.get_enquiry_type_display())
                        ws.cell(row=row_num, column=6, value=lead.remarks or "-")
                        ws.cell(row=row_num, column=7, value=lead.created_at.strftime("%d/%m/%Y %H:%M"))
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    response = HttpResponse(
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    response["Content-Disposition"] = f'attachment; filename="leads_report_{fd.strftime("%d%m%Y")}_{td.strftime("%d%m%Y")}.xlsx"'
                    wb.save(response)
                    return response
                
            except ValueError:
                messages.error(request, "Invalid date format.")
    
    return render(request, "admin/report/leads_report.html", {
        "leads": leads,
        "selected": selected,
    })

def profit_report(request):
    resorts = Resort.objects.all().order_by("resort_name")
    records = []
    employee_view = False
    employees = Employee.objects.none()
    
    selected = {"from_date": "", "to_date": "", "resort": "", "employee": ""}
    
    if request.method == "POST":
        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")
        resort_id = request.POST.get("resort")
        employee_view = request.POST.get("employee_view") == "on"
        employee_id = request.POST.get("employee")
        action = request.POST.get("action")
        
        selected = {
            "from_date": from_date or "",
            "to_date": to_date or "",
            "resort": resort_id or "",
            "employee": employee_id or "",
        }
        
        if from_date and to_date and resort_id:
            try:
                fd = datetime.strptime(from_date, "%Y-%m-%d").date()
                td = datetime.strptime(to_date, "%Y-%m-%d").date()
                
                # Get invoices
                invoice_qs = Invoice.objects.select_related(
                    "customer", "resort", "sales_person"
                ).filter(resort_id=resort_id, invoice_date__range=(fd, td))
                
                # Get vouchers
                voucher_qs = Voucher.objects.select_related(
                    "customer", "resort", "sales_person"
                ).filter(resort_id=resort_id, voucher_date__range=(fd, td))
                
                # Filter employees who have invoices or vouchers in the selected date range and resort
                if employee_view:
                    invoice_emp_ids = invoice_qs.filter(
                        sales_person__isnull=False
                    ).values_list('sales_person_id', flat=True).distinct()
                    
                    voucher_emp_ids = voucher_qs.filter(
                        sales_person__isnull=False
                    ).values_list('sales_person_id', flat=True).distinct()
                    
                    employee_ids = set(list(invoice_emp_ids) + list(voucher_emp_ids))
                    
                    employees = Employee.objects.filter(
                        id__in=employee_ids,
                        status="Active"
                    ).order_by("name")
                
                # Employee filter
                if employee_view and employee_id:
                    invoice_qs = invoice_qs.filter(sales_person_id=employee_id)
                    voucher_qs = voucher_qs.filter(sales_person_id=employee_id)
                
                # Combine invoices and vouchers into a unified list
                for inv in invoice_qs:
                    records.append({
                        'type': 'Invoice',
                        'number': inv.invoice_no,
                        'date': inv.invoice_date,
                        'customer': inv.customer.display_name if inv.customer else '-',
                        'sales_person': inv.sales_person.name if inv.sales_person else '-',
                        'resort': inv.resort.resort_name if inv.resort else '-',
                        'total': inv.total,
                        'resort_cost': inv.resort_price,
                        'profit': inv.profit,
                    })
                
                for vou in voucher_qs:
                    records.append({
                        'type': 'Voucher',
                        'number': vou.voucher_no,
                        'date': vou.voucher_date,
                        'customer': vou.customer.display_name if vou.customer else '-',
                        'sales_person': vou.sales_person.name if vou.sales_person else '-',
                        'resort': vou.resort.resort_name if vou.resort else '-',
                        'total': vou.total_amount,
                        'resort_cost': vou.resort_price,
                        'profit': vou.profit,
                    })
                
                # Sort by date descending
                records = sorted(records, key=lambda x: x['date'], reverse=True)
                
                # Excel export
                if action == "excel":
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Profit Report"
                    
                    # Header styling
                    header_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    # Headers
                    headers = ["Type", "Number", "Date", "Customer", "Resort", "Total", "Resort Cost", "Profit"]
                    if employee_view:
                        headers.insert(4, "Sales Person")
                    
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Data rows
                    for row_num, record in enumerate(records, 2):
                        ws.cell(row=row_num, column=1, value=record['type'])
                        ws.cell(row=row_num, column=2, value=record['number'])
                        ws.cell(row=row_num, column=3, value=record['date'].strftime("%d/%m/%Y"))
                        ws.cell(row=row_num, column=4, value=record['customer'])
                        
                        if employee_view:
                            ws.cell(row=row_num, column=5, value=record['sales_person'])
                            ws.cell(row=row_num, column=6, value=record['resort'])
                            ws.cell(row=row_num, column=7, value=float(record['total']))
                            ws.cell(row=row_num, column=8, value=float(record['resort_cost']))
                            ws.cell(row=row_num, column=9, value=float(record['profit']))
                        else:
                            ws.cell(row=row_num, column=5, value=record['resort'])
                            ws.cell(row=row_num, column=6, value=float(record['total']))
                            ws.cell(row=row_num, column=7, value=float(record['resort_cost']))
                            ws.cell(row=row_num, column=8, value=float(record['profit']))
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    response = HttpResponse(
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    response["Content-Disposition"] = f'attachment; filename="profit_report_{fd.strftime("%d%m%Y")}_{td.strftime("%d%m%Y")}.xlsx"'
                    wb.save(response)
                    return response
                
            except ValueError:
                messages.error(request, "Invalid date format.")
    
    return render(request, "admin/report/profit_report.html", {
        "resorts": resorts,
        "records": records,
        "employees": employees,
        "employee_view": employee_view,
        "selected": selected,
    })

# DESTINATION VIEWS
def destination_list(request):
    category = request.GET.get('cat', 'Domestic')
    search_query = request.GET.get('search', '').strip()
    
    destinations = Destination.objects.filter(category=category)
    
    if search_query:
        destinations = destinations.filter(
            Q(name__icontains=search_query) |
            Q(country__icontains=search_query)
        )
    
    destinations = destinations.order_by('-created_at')
    domestic_count = Destination.objects.filter(category='Domestic').count()
    international_count = Destination.objects.filter(category='International').count()
    context = {
        'destinations': destinations,
        'selected_category': category,
        'domestic_count': domestic_count,
        'international_count': international_count,
        'search_query': search_query,
    }
    return render(request, 'admin/destination/destination.html', context)

def add_destination(request):
    # Get category from URL parameter (from travel packages page)
    default_category = request.GET.get('cat', 'Domestic')
    
    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        country = request.POST.get('country', '').strip()
        category = request.POST.get('category')
        description = request.POST.get('description', '').strip()
        
        if not name or not country:
            messages.error(request, "Name and country are required.")
            return render(request, 'admin/destination/add_destination.html', {'default_category': default_category})
        
        # Validate name length
        if len(name) < 2 or len(name) > 100:
            messages.error(request, "Destination name must be between 2 and 100 characters.")
            return render(request, 'admin/destination/add_destination.html', {'default_category': default_category})
        
        # Validate country length
        if len(country) < 2 or len(country) > 100:
            messages.error(request, "Country name must be between 2 and 100 characters.")
            return render(request, 'admin/destination/add_destination.html', {'default_category': default_category})
        
        # Check for duplicate destination
        if Destination.objects.filter(name__iexact=name, country__iexact=country).exists():
            messages.error(request, "This destination already exists.")
            return render(request, 'admin/destination/add_destination.html', {'default_category': default_category})
        
        # Validate description length if provided
        if description and len(description) > 500:
            messages.error(request, "Description must not exceed 500 characters.")
            return render(request, 'admin/destination/add_destination.html', {'default_category': default_category})

        packages_start_from = request.POST.get('packages_start_from', '').strip()
        Destination.objects.create(
            name=name,
            country=country,
            category=category,
            description=description,
            is_popular=request.POST.get('is_popular') == 'on',
            packages_start_from=packages_start_from if packages_start_from else None,
            image=request.FILES.get('image'),
            map_image=request.FILES.get('map_image')
        )
        messages.success(request, "Destination added successfully!")
        
        # Redirect back to destinations with the category
        url = reverse('admin_panel:destinations')
        return redirect(f'{url}?cat={category}')
    
    context = {'default_category': default_category}
    return render(request, 'admin/destination/add_destination.html', context)

def view_destination(request, destination_id):
    destination = get_object_or_404(Destination, id=destination_id)
    return render(request, 'admin/destination/view_destination.html', {'destination': destination})

def edit_destination(request, destination_id):
    destination = get_object_or_404(Destination, id=destination_id)
    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        country = request.POST.get('country', '').strip()
        category = request.POST.get('category')
        if not name or not country:
            messages.error(request, "Name and country are required.")
            return render(request, 'admin/destination/edit_destination.html', {'destination': destination})
        destination.name = name
        destination.country = country
        destination.category = category
        destination.description = request.POST.get('description')
        destination.is_popular = request.POST.get('is_popular') == 'on'
        packages_start_from = request.POST.get('packages_start_from', '').strip()
        destination.packages_start_from = packages_start_from if packages_start_from else None
        if request.FILES.get('image'):
            destination.image = request.FILES.get('image')
        if request.FILES.get('map_image'):
            destination.map_image = request.FILES.get('map_image')
        destination.save()
        messages.success(request, "Destination updated successfully!")
        return redirect(f"{reverse('admin_panel:destinations')}?cat={category}")
    return render(request, 'admin/destination/edit_destination.html', {'destination': destination})

def delete_destination(request, destination_id):
    destination = get_object_or_404(Destination, id=destination_id)
    destination.delete()
    messages.success(request, "Destination deleted successfully!")
    return redirect('admin_panel:destinations')

# FEEDBACK MANAGEMENT
def feedback_list(request):
    feedbacks = Feedback.objects.all().order_by('-created_at')

    search_query = request.GET.get('search', '').strip()
    if search_query:
        feedbacks = feedbacks.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(mobile_number__icontains=search_query)
        )

    rating_filter = request.GET.get('rating', '').strip()
    if rating_filter:
        feedbacks = feedbacks.filter(rating=rating_filter)

    type_filter = request.GET.get('feedback_type', '').strip()
    if type_filter:
        feedbacks = feedbacks.filter(feedback_type=type_filter)

    featured_count = Feedback.objects.filter(featured=True).count()

    context = {
        'feedbacks': feedbacks,
        'search_query': search_query,
        'rating_filter': rating_filter,
        'type_filter': type_filter,
        'total_feedbacks': Feedback.objects.count(),
        'featured_count': featured_count,
        'feedback_type_choices': Feedback.FEEDBACK_TYPE_CHOICES,
    }
    return render(request, 'admin/feedback/feedback.html', context)
def add_feedback(request):
    if request.method == 'POST':
        try:
            name = (request.POST.get('name') or '').strip()
            email = (request.POST.get('email') or '').strip()
            mobile_number = (request.POST.get('mobile_number') or '').strip()
            feedback_type = (request.POST.get('feedback_type') or '').strip()
            rating = (request.POST.get('rating') or '').strip()
            feedback_text = (request.POST.get('feedback') or '').strip()

            if not all([name, email, rating, feedback_text, feedback_type]):
                messages.error(request, "Name, Email, Feedback Type, Rating and Feedback are required.")
                return render(request, 'admin/feedback/add_feedback.html', {
                    'feedback_type_choices': Feedback.FEEDBACK_TYPE_CHOICES
                })

            featured_flag = request.POST.get('featured') == '1'

            feedback_obj = Feedback.objects.create(
                name=name,
                email=email,
                mobile_number=mobile_number,
                feedback_type=feedback_type,
                rating=int(rating),
                feedback=feedback_text,
                featured=featured_flag
            )

            from admin_panel.models import FeedbackImage
            images = request.FILES.getlist('images')
            for img in images:
                FeedbackImage.objects.create(feedback=feedback_obj, image=img)

            messages.success(request, "Feedback added successfully!")
            return redirect('feedback:feedback_list')

        except Exception as e:
            messages.error(request, f"Error adding feedback: {str(e)}")
            return render(request, 'admin/feedback/add_feedback.html', {
                'feedback_type_choices': Feedback.FEEDBACK_TYPE_CHOICES
            })

    return render(request, 'admin/feedback/add_feedback.html', {
        'feedback_type_choices': Feedback.FEEDBACK_TYPE_CHOICES
    })
def view_feedback(request, feedback_id):
    feedback = get_object_or_404(Feedback, id=feedback_id)
    return render(request, 'admin/feedback/view_feedback.html', {'feedback': feedback})

def delete_feedback(request, feedback_id):
    if request.method != 'POST':
        return redirect('feedback:feedback_list')
    
    feedback = get_object_or_404(Feedback, id=feedback_id)
    feedback.delete()
    messages.success(request, 'Feedback deleted successfully!')
    return redirect('feedback:feedback_list')


def toggle_featured_feedback(request, feedback_id):
    """
    Toggle the featured flag for a feedback entry. When marked, it appears
    on the public homepage "Hear from them" card. Clicking again unmarks it.
    """
    if request.method != 'POST':
        # for safety require POST; could also allow GET but keep pattern consistent
        return redirect('feedback:feedback_list')

    feedback = get_object_or_404(Feedback, id=feedback_id)
    feedback.featured = not feedback.featured
    feedback.save()
    if feedback.featured:
        messages.success(request, 'Feedback marked as featured.')
    else:
        messages.info(request, 'Feedback unmarked as featured.')
    return redirect('feedback:feedback_list')